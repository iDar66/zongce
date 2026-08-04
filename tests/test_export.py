from pathlib import Path
import pandas as pd
import openpyxl
from zongce.predict import Prediction
from zongce.classify import Mode
from zongce.export import export_excel, organize_files
from zongce.grades import GradeSummary
from zongce.score import PanelScore, ScoreReport


def _score_report() -> ScoreReport:
    grades = GradeSummary("2025-2026", (), 53.0, 4559.0, 86.02, (), (1, 2))
    return ScoreReport(
        grades,
        PanelScore(3.0, 70.0, None, 3.0, 73.0, None, "不适用", 1),
        PanelScore(5.0, 68.816, 20.0, 4.0, 72.816, 25.0, "估算", 1),
        PanelScore(1.0, 60.0, 40.0, 0.8, 60.8, 50.0, "估算", 0),
        70.23,
    )

def _mk_predictions():
    return [
        Prediction("参赛.pdf", "文体", Mode.COUNT, 4.0, 4, "每场次加1文体分", "自动", ""),
        Prediction("后勤.pdf", "品德", Mode.COUNT, 1.0, 2, "每次加0.5品德分", "自动", ""),
        Prediction("数模.pdf", "学业", Mode.GRADE, None, None, "三等奖", "待确认", "级别待认定"),
        Prediction("新年.pdf", "品德", Mode.RULE_REF, None, None, "工作人员名单", "待确认", ""),
    ]

def test_export_excel_has_columns_and_rows(tmp_path):
    out = export_excel(_mk_predictions(), tmp_path / "综测加分明细.xlsx")
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header[:7] == ["类别", "项目", "级别/明细", "细则依据", "加分", "认定状态", "备注"]
    texts = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "参赛" in texts and "数模" in texts   # 项目列取 stem
    assert "▶" in texts  # 至少一个小计行

def test_organize_files_copies_into_panels(tmp_path):
    src_a = tmp_path / "参赛.pdf"; src_a.write_text("x")
    src_b = tmp_path / "数模.pdf"; src_b.write_text("y")
    preds = [
        Prediction(str(src_a), "文体", Mode.COUNT, 4.0, 4, "", "自动", ""),
        Prediction(str(src_b), "学业", Mode.GRADE, None, None, "", "待确认", ""),
    ]
    out = organize_files(preds, tmp_path / "归类")
    assert (out / "文体" / "参赛.pdf").exists()
    assert (out / "学业" / "数模.pdf").exists()

def test_organize_files_renames_duplicates(tmp_path):
    d1 = tmp_path / "d1"; d1.mkdir(); src = d1 / "a.pdf"; src.write_text("x")
    d2 = tmp_path / "d2"; d2.mkdir(); src2 = d2 / "a.pdf"; src2.write_text("y")
    preds = [
        Prediction(str(src), "文体", Mode.COUNT, 1.0, 1, "", "自动", ""),
        Prediction(str(src2), "文体", Mode.COUNT, 1.0, 1, "", "自动", ""),
    ]
    out = organize_files(preds, tmp_path / "归类2")
    names = sorted(p.name for p in (out / "文体").iterdir())
    assert names == ["a.pdf", "a_1.pdf"]


def test_export_excel_adds_score_sheet_when_report_is_given(tmp_path):
    out = export_excel(_mk_predictions(), tmp_path / "评分.xlsx", score_report=_score_report())

    workbook = openpyxl.load_workbook(out)

    assert workbook.sheetnames == ["综测加分明细", "综测评分预测"]
    score_sheet = workbook["综测评分预测"]
    headers = [cell.value for cell in score_sheet[1]]
    values = " ".join(str(cell.value) for row in score_sheet.iter_rows() for cell in row if cell.value is not None)
    assert "班级最高 raw 来源" in headers
    assert "估算" in values
