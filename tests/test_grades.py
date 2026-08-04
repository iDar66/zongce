# -*- coding: utf-8 -*-
"""成绩 Excel 读取测试。"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
import pytest


HEADERS = ["学年", "学期", "课程代码", "课程名称", "学分", "成绩分项", "成绩"]


def _write_grade_book(
    path: Path,
    semester: int,
    rows: list[tuple[str, str, float, str, float]],
    academic_year: str = "2025-2026",
) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    for course_code, course_name, credits, item, score in rows:
        sheet.append([academic_year, semester, course_code, course_name, credits, item, score])
    workbook.save(path)
    return path


def test_read_grade_files_uses_total_score_rows_for_weighted_average(tmp_path):
    from zongce.grades import read_grade_files

    upper = _write_grade_book(
        tmp_path / "upper.xlsx",
        1,
        [("A001", "课程 A", 3.0, "平时(30%)", 100.0), ("A001", "课程 A", 3.0, "总评", 90.0)],
    )
    lower = _write_grade_book(
        tmp_path / "lower.xlsx",
        2,
        [("B001", "课程 B", 1.0, "总评", 80.0)],
    )

    summary = read_grade_files([upper, lower])

    assert summary.course_count == 2
    assert summary.total_credits == 4.0
    assert summary.weighted_average == pytest.approx(87.5)


def test_read_grade_files_rejects_missing_second_semester(tmp_path):
    from zongce.grades import GradeInputError, read_grade_files

    first = _write_grade_book(
        tmp_path / "first.xlsx",
        1,
        [("A001", "课程 A", 3.0, "总评", 90.0)],
    )
    duplicate = _write_grade_book(
        tmp_path / "duplicate.xlsx",
        1,
        [("B001", "课程 B", 1.0, "总评", 80.0)],
    )

    with pytest.raises(GradeInputError, match="学期"):
        read_grade_files([first, duplicate])


def test_read_grade_files_rejects_different_academic_years(tmp_path):
    from zongce.grades import GradeInputError, read_grade_files

    upper = _write_grade_book(
        tmp_path / "upper.xlsx",
        1,
        [("A001", "课程 A", 3.0, "总评", 90.0)],
    )
    lower = _write_grade_book(
        tmp_path / "lower.xlsx",
        2,
        [("B001", "课程 B", 1.0, "总评", 80.0)],
        academic_year="2026-2027",
    )

    with pytest.raises(GradeInputError, match="学年"):
        read_grade_files([upper, lower])


def test_read_grade_files_rejects_duplicate_total_score_for_course(tmp_path):
    from zongce.grades import GradeInputError, read_grade_files

    upper = _write_grade_book(
        tmp_path / "upper.xlsx",
        1,
        [("A001", "课程 A", 3.0, "总评", 90.0), ("A001", "课程 A", 3.0, "总评", 90.0)],
    )
    lower = _write_grade_book(
        tmp_path / "lower.xlsx",
        2,
        [("B001", "课程 B", 1.0, "总评", 80.0)],
    )

    with pytest.raises(GradeInputError, match="重复"):
        read_grade_files([upper, lower])


def test_read_grade_files_rejects_non_numeric_total_score(tmp_path):
    from zongce.grades import GradeInputError, read_grade_files

    upper = _write_grade_book(
        tmp_path / "upper.xlsx",
        1,
        [("A001", "课程 A", 3.0, "总评", "优秀")],
    )
    lower = _write_grade_book(
        tmp_path / "lower.xlsx",
        2,
        [("B001", "课程 B", 1.0, "总评", 80.0)],
    )

    with pytest.raises(GradeInputError, match="数值"):
        read_grade_files([upper, lower])


def test_read_grade_files_rejects_missing_required_column(tmp_path):
    from zongce.grades import GradeInputError, read_grade_files

    upper = _write_grade_book(
        tmp_path / "upper.xlsx",
        1,
        [("A001", "课程 A", 3.0, "总评", 90.0)],
    )
    lower = _write_grade_book(
        tmp_path / "lower.xlsx",
        2,
        [("B001", "课程 B", 1.0, "总评", 80.0)],
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["学年", "学期", "课程代码", "课程名称", "学分", "成绩分项"])
    sheet.append(["2025-2026", 2, "B001", "课程 B", 1.0, "总评"])
    workbook.save(lower)

    with pytest.raises(GradeInputError, match="缺少列"):
        read_grade_files([upper, lower])


def test_read_grade_files_rejects_missing_file(tmp_path):
    from zongce.grades import GradeInputError, read_grade_files

    with pytest.raises(GradeInputError, match="不存在"):
        read_grade_files([tmp_path / "missing.xlsx"])


def test_read_grade_files_rejects_zero_total_credits(tmp_path):
    from zongce.grades import GradeInputError, read_grade_files

    # 所有总评行学分均为 0（校验只拒 <0，允许 0）→ 不应 ZeroDivisionError，应 GradeInputError
    upper = _write_grade_book(tmp_path / "upper.xlsx", 1, [("A001", "课程 A", 0.0, "总评", 90.0)])
    lower = _write_grade_book(tmp_path / "lower.xlsx", 2, [("B001", "课程 B", 0.0, "总评", 80.0)])

    with pytest.raises(GradeInputError, match="学分"):
        read_grade_files([upper, lower])


def test_read_grade_files_rejects_non_numeric_semester(tmp_path):
    from zongce.grades import GradeInputError, read_grade_files

    # 学期单元格为非数值字符串 → 应转 GradeInputError，而非裸 ValueError traceback
    upper = tmp_path / "upper.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    sheet.append(["2025-2026", "第一学期", "A001", "课程 A", 3.0, "总评", 90.0])
    workbook.save(upper)
    lower = _write_grade_book(tmp_path / "lower.xlsx", 2, [("B001", "课程 B", 1.0, "总评", 80.0)])

    with pytest.raises(GradeInputError, match="数值"):
        read_grade_files([upper, lower])
